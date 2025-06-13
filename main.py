import os
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def extract_sql_components(file_path):
    """Extract databases, tables, and columns from an SQL file."""
    with open(file_path, 'r', encoding='utf-8') as file:
        content = file.read()
    
    # Remove comments and standardize whitespace
    content = re.sub(r'--.*?\n|/\*[\s\S]*?\*/|\n|\t', ' ', content)
    content = re.sub(r'\s+', ' ', content).strip()
    
    # Patterns to extract components
    # Extract database.table.column patterns
    db_table_column_pattern = r'(\w+)\.(\w+)\.(\w+)'
    db_table_column_matches = re.findall(db_table_column_pattern, content)
    
    # Extract table.column patterns
    table_column_pattern = r'(\w+)\.(\w+)'
    table_column_matches = re.findall(table_column_pattern, content)
    
    # Extract FROM clauses to identify tables
    from_pattern = r'FROM\s+(\w+\.)?(\w+)'
    from_matches = re.findall(from_pattern, content, re.IGNORECASE)
    
    # Extract JOIN clauses to identify tables
    join_pattern = r'JOIN\s+(\w+\.)?(\w+)'
    join_matches = re.findall(join_pattern, content, re.IGNORECASE)
    
    # Extract SELECT clauses to identify columns
    select_pattern = r'SELECT\s+(.*?)\s+FROM'
    select_matches = re.findall(select_pattern, content, re.IGNORECASE)
    
    # Process the columns from SELECT clause
    columns = []
    if select_matches:
        select_columns = select_matches[0].split(',')
        for col in select_columns:
            col = col.strip()
            # Extract column name (handle cases like 'table.column as alias')
            col_match = re.search(r'(?:(\w+)\.)?(\w+)(?:\s+[aA][sS]\s+(\w+))?', col)
            if col_match:
                # Get column name, ignoring alias
                column_name = col_match.group(2)
                if column_name and column_name.upper() not in ('SELECT', 'FROM', 'WHERE', 'JOIN'):
                    columns.append(column_name)
    
    # Combine the results
    result = {
        'database': set(),
        'table': set(),
        'column': set(),
        'relationships': []  # To store db.table.column relationships
    }
    
    # Add database.table.column matches
    for db, table, column in db_table_column_matches:
        result['database'].add(db)
        result['table'].add(table)
        result['column'].add(column)
        result['relationships'].append((db, table, column))
    
    # Add table.column matches
    for table, column in table_column_matches:
        # Skip if the table is actually a database (from db.table.column)
        if table not in result['database']:
            result['table'].add(table)
            result['column'].add(column)
            # Add with empty database
            result['relationships'].append(('', table, column))
    
    # Add tables from FROM clauses
    for db_prefix, table in from_matches:
        if table.upper() not in ('SELECT', 'FROM', 'WHERE', 'JOIN'):
            result['table'].add(table)
            if db_prefix:
                db = db_prefix.rstrip('.')
                result['database'].add(db)
    
    # Add tables from JOIN clauses
    for db_prefix, table in join_matches:
        if table.upper() not in ('SELECT', 'FROM', 'WHERE', 'JOIN'):
            result['table'].add(table)
            if db_prefix:
                db = db_prefix.rstrip('.')
                result['database'].add(db)
    
    # Add columns from SELECT clause
    for column in columns:
        result['column'].add(column)
    
    return result

def process_sql_folder(folder_path):
    """Process all SQL files in a folder and extract components."""
    all_components = {
        'database': set(),
        'table': set(),
        'column': set(),
        'relationships': []
    }
    
    # Get all SQL files in the folder
    sql_files = [f for f in os.listdir(folder_path) if f.endswith('.sql')]
    
    for sql_file in sql_files:
        file_path = os.path.join(folder_path, sql_file)
        file_components = extract_sql_components(file_path)
        
        # Merge the results
        all_components['database'].update(file_components['database'])
        all_components['table'].update(file_components['table'])
        all_components['column'].update(file_components['column'])
        all_components['relationships'].extend(file_components['relationships'])
    
    return all_components

def create_excel_data_dictionary(components, output_file):
    """Create an Excel data dictionary based on the extracted components."""
    # Create a new Excel workbook
    workbook = Workbook()
    
    # Create sheets
    overview_sheet = workbook.active
    overview_sheet.title = "Overview"
    database_sheet = workbook.create_sheet(title="Databases")
    table_sheet = workbook.create_sheet(title="Tables")
    column_sheet = workbook.create_sheet(title="Columns")
    relationship_sheet = workbook.create_sheet(title="Relationships")
    
    # Style definitions
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Overview sheet
    overview_sheet["A1"] = "Data Dictionary Overview"
    overview_sheet["A1"].font = Font(bold=True, size=16)
    overview_sheet["A3"] = "Total Databases:"
    overview_sheet["B3"] = len(components['database'])
    overview_sheet["A4"] = "Total Tables:"
    overview_sheet["B4"] = len(components['table'])
    overview_sheet["A5"] = "Total Columns:"
    overview_sheet["B5"] = len(components['column'])
    overview_sheet["A6"] = "Total Relationships:"
    overview_sheet["B6"] = len(components['relationships'])
    
    # Databases sheet
    database_sheet["A1"] = "Database Name"
    database_sheet["B1"] = "Description"
    database_sheet["A1"].font = header_font
    database_sheet["B1"].font = header_font
    database_sheet["A1"].fill = header_fill
    database_sheet["B1"].fill = header_fill
    
    row = 2
    for database in sorted(components['database']):
        database_sheet[f"A{row}"] = database
        database_sheet[f"B{row}"] = ""  # Description to be filled manually
        row += 1
    
    # Tables sheet
    table_sheet["A1"] = "Table Name"
    table_sheet["B1"] = "Database"
    table_sheet["C1"] = "Description"
    table_sheet["A1"].font = header_font
    table_sheet["B1"].font = header_font
    table_sheet["C1"].font = header_font
    table_sheet["A1"].fill = header_fill
    table_sheet["B1"].fill = header_fill
    table_sheet["C1"].fill = header_fill
    
    # Get unique table-database pairs
    table_db_pairs = set()
    for db, table, _ in components['relationships']:
        if db and table:
            table_db_pairs.add((table, db))
    
    # Add tables without explicit database relation
    standalone_tables = components['table'] - {table for table, _ in table_db_pairs}
    
    row = 2
    # First add tables with known database relations
    for table, db in sorted(table_db_pairs):
        table_sheet[f"A{row}"] = table
        table_sheet[f"B{row}"] = db
        table_sheet[f"C{row}"] = ""  # Description to be filled manually
        row += 1
    
    # Then add standalone tables
    for table in sorted(standalone_tables):
        table_sheet[f"A{row}"] = table
        table_sheet[f"B{row}"] = ""
        table_sheet[f"C{row}"] = ""
        row += 1
    
    # Columns sheet
    column_sheet["A1"] = "Column Name"
    column_sheet["B1"] = "Table"
    column_sheet["C1"] = "Database"
    column_sheet["D1"] = "Data Type"
    column_sheet["E1"] = "Description"
    for cell in ["A1", "B1", "C1", "D1", "E1"]:
        column_sheet[cell].font = header_font
        column_sheet[cell].fill = header_fill
    
    # Get unique column-table-database relationships
    col_relationships = set()
    for db, table, col in components['relationships']:
        col_relationships.add((col, table, db))
    
    # Add columns without explicit table relation
    standalone_columns = components['column'] - {col for col, _, _ in col_relationships}
    
    row = 2
    # First add columns with known relationships
    for col, table, db in sorted(col_relationships):
        column_sheet[f"A{row}"] = col
        column_sheet[f"B{row}"] = table
        column_sheet[f"C{row}"] = db
        column_sheet[f"D{row}"] = ""  # Data type to be filled manually
        column_sheet[f"E{row}"] = ""  # Description to be filled manually
        row += 1
    
    # Then add standalone columns
    for col in sorted(standalone_columns):
        column_sheet[f"A{row}"] = col
        column_sheet[f"B{row}"] = ""
        column_sheet[f"C{row}"] = ""
        column_sheet[f"D{row}"] = ""
        column_sheet[f"E{row}"] = ""
        row += 1
    
    # Relationships sheet
    relationship_sheet["A1"] = "Database"
    relationship_sheet["B1"] = "Table"
    relationship_sheet["C1"] = "Column"
    for cell in ["A1", "B1", "C1"]:
        relationship_sheet[cell].font = header_font
        relationship_sheet[cell].fill = header_fill
    
    row = 2
    for db, table, col in sorted(components['relationships']):
        relationship_sheet[f"A{row}"] = db
        relationship_sheet[f"B{row}"] = table
        relationship_sheet[f"C{row}"] = col
        row += 1
    
    # Auto-adjust column widths
    for sheet in [overview_sheet, database_sheet, table_sheet, column_sheet, relationship_sheet]:
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    workbook.save(output_file)
    print(f"Data dictionary saved to {output_file}")

def main():
    # Get folder path from user
    folder_path = input("Enter the path to the folder containing SQL files: ")
    output_file = input("Enter the path for the output Excel file (default: data_dictionary.xlsx): ")
    
    if not output_file:
        output_file = "data_dictionary.xlsx"
    
    # Check if folder exists
    if not os.path.isdir(folder_path):
        print(f"Error: Folder '{folder_path}' does not exist.")
        return
    
    # Process SQL files
    print(f"Processing SQL files in '{folder_path}'...")
    components = process_sql_folder(folder_path)
    
    # Create Excel data dictionary
    create_excel_data_dictionary(components, output_file)

if __name__ == "__main__":
    main()